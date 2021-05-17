import openpyxl
inv_file = openpyxl.load_workbook("existigfile.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_persupplier = {}
count_inventunder10 = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inv_count = product_list.cell(product_row, 2).value
    prod_id = product_list.cell(product_row, 1).value
    price = product_list.cell(product_row, 3).value
    inventcomplete_price = product_list.cell(product_row, 5)

    print(supplier_name)
    print(inv_count)


    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a supplier")
        products_per_supplier[supplier_name] = 1

    if supplier_name in total_value_persupplier:
        current_count = total_value_persupplier[supplier_name]
        total_value_persupplier[supplier_name] = current_count + inv_count
    else:
        print("adding a count")
        total_value_persupplier[supplier_name] = inv_count



    if (inv_count < 10):
        count_inventunder10[prod_id] = inv_count
    inventcomplete_price.value = inv_count * price

print(products_per_supplier)
print(total_value_persupplier)
print(count_inventunder10)
print(inventcomplete_price)

inv_file.save("newfile.xlsx")